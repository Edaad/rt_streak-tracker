import pandas as pd
import os
import argparse
from datetime import datetime
import random
import openpyxl
from referral_system import ReferralSystem
import config


class PokerStreakTracker:
    def __init__(self, master_file_path, hands_threshold=100, history_file_path=None):
        """
        Initialize the tracker with local file paths and configuration.

        Args:
            master_file_path: Path to master Excel file
            hands_threshold: Minimum hands required to increment streak (default: 100)
            history_file_path: Path to history Excel file (default: derived from master path)
        """
        self.master_file_path = master_file_path
        self.hands_threshold = hands_threshold

        # Set default history path if not provided
        if history_file_path is None:
            master_dir = os.path.dirname(master_file_path)
            master_name = os.path.splitext(os.path.basename(master_file_path))[0]
            self.history_file_path = os.path.join(
                master_dir, f"{master_name}_history.xlsx"
            )
        else:
            self.history_file_path = history_file_path

        self.wheel_milestones = [
            7,
            14,
            21,
            28,
            35,
            42,
            49,
            56,
            63,
            70,
        ]  # Multiples of 7

        # Create master file if it doesn't exist
        if not os.path.exists(master_file_path):
            print(
                f"Master file not found. Creating new master file at {master_file_path}"
            )
            df_master = pd.DataFrame(columns=["Username", "Streak"])
            df_master.to_excel(master_file_path, index=False)

        # Create history file if it doesn't exist
        if not os.path.exists(self.history_file_path):
            print(
                f"History file not found. Creating new history file at {self.history_file_path}"
            )
            df_history = pd.DataFrame(
                columns=[
                    "Username",
                    "LastUpdate",
                    "UpdateDate",
                    "CurrentStreak",
                    "HighestStreak",
                ]
            )
            df_history.to_excel(self.history_file_path, index=False)

    def _generate_unique_wheel_times(self, count):
        """Generate unique time slots between 7-9 PM in 10-minute increments."""
        # Available time slots (7:00 PM to 8:50 PM in 10-minute increments)
        available_slots = []
        for hour in [19, 20]:
            for minute in [0, 10, 20, 30, 40, 50]:
                available_slots.append(f"{hour}:{minute:02d} PM")

        # Ensure we have enough slots
        if count > len(available_slots):
            # If we need more slots than available, add some additional slots
            for hour in [21]:  # Extend to 9 PM if needed
                for minute in [0, 10, 20, 30, 40, 50]:
                    available_slots.append(f"{hour}:{minute:02d} PM")

            # If we still need more slots, add additional hours
            if count > len(available_slots):
                for hour in [18, 22]:  # Add 6 PM and 10 PM if needed
                    for minute in [0, 10, 20, 30, 40, 50]:
                        available_slots.append(f"{hour}:{minute:02d} PM")

        # Randomly select the required number of unique slots
        selected_slots = random.sample(
            available_slots, min(count, len(available_slots))
        )

        # If we have more winners than available slots, reuse some slots
        while len(selected_slots) < count:
            extra_slot = random.choice(available_slots)
            selected_slots.append(f"{extra_slot} (overflow)")

        return selected_slots

    def _get_wheel_number(self, streak):
        """Get the wheel number based on streak milestone."""
        for i, milestone in enumerate(self.wheel_milestones):
            if streak == milestone:
                return i + 1
        return None

    def _extract_player_data_from_complex_excel(self, file_path):
        """
        Extract player data from complex Excel format.

        Specifically:
        - Sheet name: "Member Statistics"
        - Column J (index 10): Player usernames
        - Column EV: Hands played
        """
        try:
            from openpyxl.utils import column_index_from_string

            # Load the workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            # Check if the required sheet exists
            if "Member Statistics" not in workbook.sheetnames:
                print(
                    f"Error: Required sheet 'Member Statistics' not found in {file_path}"
                )
                return None

            sheet = workbook["Member Statistics"]

            # Define column letters
            username_column = "J"
            hands_column = "EV"

            # Get column indices
            username_col_idx = column_index_from_string(username_column)
            hands_col_idx = column_index_from_string(hands_column)

            # Extract player data
            player_data = []

            # Start from row 7 (which is index 6 in zero-based counting)
            row_index = 7

            while True:
                username_cell = sheet.cell(
                    row=row_index, column=username_col_idx
                )  # Column J
                hands_cell = sheet.cell(
                    row=row_index, column=hands_col_idx
                )  # Column EV

                # Break if we've reached the end of data
                if username_cell.value is None:
                    break

                username = str(username_cell.value).strip()

                # Skip rows with empty usernames
                if not username:
                    row_index += 1
                    continue

                # Handle different data types for hands value
                hands_value = hands_cell.value
                if hands_value is None:
                    hands = 0
                elif isinstance(hands_value, (int, float)):
                    hands = int(hands_value)
                else:
                    try:
                        hands = int(str(hands_value).strip())
                    except (ValueError, TypeError):
                        print(
                            f"Warning: Invalid hands value for {username}: '{hands_value}'. Setting to 0."
                        )
                        hands = 0

                player_data.append({"Username": username, "Hands": hands})

                row_index += 1

            print(f"Extracted data for {len(player_data)} players")
            return player_data

        except Exception as e:
            print(f"Error extracting data from complex Excel file: {str(e)}")
            return None

    def _update_player_history(self, username, status_message, current_streak=0):
        """
        Update the player history with their latest status.

        Args:
            username: Player username
            status_message: Status update message
            current_streak: Current streak value (0 if inactive)
        """
        try:
            # Load the history file
            if os.path.exists(self.history_file_path):
                df_history = pd.read_excel(self.history_file_path)
            else:
                df_history = pd.DataFrame(
                    columns=[
                        "Username",
                        "LastUpdate",
                        "UpdateDate",
                        "CurrentStreak",
                        "HighestStreak",
                    ]
                )

            # Get today's date formatted for display
            today_date = datetime.now().strftime("%B %d, %Y")

            # Check if player already exists in history
            player_exists = username in df_history["Username"].values

            if player_exists:
                # Update existing player
                player_idx = df_history[df_history["Username"] == username].index[0]
                highest_streak = df_history.at[player_idx, "HighestStreak"]

                # Update highest streak if current is higher
                if current_streak > highest_streak:
                    highest_streak = current_streak

                # Update the player's record
                df_history.at[player_idx, "LastUpdate"] = status_message
                df_history.at[player_idx, "UpdateDate"] = today_date
                df_history.at[player_idx, "CurrentStreak"] = current_streak
                df_history.at[player_idx, "HighestStreak"] = highest_streak
            else:
                # Add new player
                new_record = {
                    "Username": username,
                    "LastUpdate": status_message,
                    "UpdateDate": today_date,
                    "CurrentStreak": current_streak,
                    "HighestStreak": current_streak,
                }
                df_history = pd.concat(
                    [df_history, pd.DataFrame([new_record])], ignore_index=True
                )

            # Save the history file
            df_history.to_excel(self.history_file_path, index=False)
            return True

        except Exception as e:
            print(f"Error updating history for player {username}: {str(e)}")
            return False

    def revive_player_streak(self, username, new_streak):
        """
        Revive or set a specific streak for a player (admin function).

        Args:
            username: Player username
            new_streak: Streak value to set

        Returns:
            Tuple of (success, message)
        """
        try:
            # Ensure streak is a positive number
            try:
                new_streak = int(new_streak)
                if new_streak < 1:
                    return False, "Streak value must be positive"
            except ValueError:
                return False, "Invalid streak value"

            # Check if master file exists
            if not os.path.exists(self.master_file_path):
                return False, f"Master file not found at {self.master_file_path}"

            # Read master file
            df_master = pd.read_excel(self.master_file_path)

            # Check if player exists in master file
            player_rows = df_master[df_master["Username"] == username]
            player_exists = len(player_rows) > 0

            if player_exists:
                # Get current streak
                player_index = player_rows.index[0]
                current_streak = df_master.at[player_index, "Streak"]

                # If current streak is higher, confirm overwrite
                if current_streak > new_streak:
                    return (
                        "confirm",
                        f"Player '{username}' currently has a higher streak ({current_streak}). Are you sure you want to set it to {new_streak}?",
                    )

                # Update streak
                df_master.at[player_index, "Streak"] = new_streak
            else:
                # Add new player
                new_player = {"Username": username, "Streak": new_streak}
                df_master = pd.concat(
                    [df_master, pd.DataFrame([new_player])], ignore_index=True
                )

            # Update player history
            status_message = f"{new_streak} day streak revived!"
            self._update_player_history(username, status_message, new_streak)

            # Save updated master file
            df_master.to_excel(self.master_file_path, index=False)

            return True, f"Successfully set {username}'s streak to {new_streak} days"

        except Exception as e:
            return False, f"Error reviving player streak: {str(e)}"

    def confirm_revive_streak(self, username, new_streak):
        """
        Confirm and execute streak revival after warning about overwriting higher streak.
        """
        try:
            # Read master file
            df_master = pd.read_excel(self.master_file_path)

            # Update player's streak
            player_rows = df_master[df_master["Username"] == username]
            if len(player_rows) > 0:
                player_index = player_rows.index[0]
                df_master.at[player_index, "Streak"] = new_streak
            else:
                # Add new player (shouldn't happen in confirmation flow but just in case)
                new_player = {"Username": username, "Streak": new_streak}
                df_master = pd.concat(
                    [df_master, pd.DataFrame([new_player])], ignore_index=True
                )

            # Update player history
            status_message = f"{new_streak} day streak revived!"
            self._update_player_history(username, status_message, new_streak)

            # Save updated master file
            df_master.to_excel(self.master_file_path, index=False)

            return (
                True,
                f"Successfully set {username}'s streak to {new_streak} days (overwriting higher streak)",
            )

        except Exception as e:
            return False, f"Error confirming streak revival: {str(e)}"

    def lookup_player(self, username):
        """
        Look up a player's status in the history database.

        Args:
            username: Player username to look up

        Returns:
            Player status dictionary or None if not found
        """
        try:
            # Check if history file exists
            if not os.path.exists(self.history_file_path):
                print(f"History file not found at {self.history_file_path}")
                return None

            # Load history file
            df_history = pd.read_excel(self.history_file_path)

            # Look up player
            player_rows = df_history[df_history["Username"] == username]

            if len(player_rows) == 0:
                return None

            # Get player data
            player_data = player_rows.iloc[0].to_dict()
            return player_data

        except Exception as e:
            print(f"Error looking up player {username}: {str(e)}")
            return None

    def process_daily_data(self, daily_file_path):
        """
        Process the daily player data and update streaks in master file.

        Args:
            daily_file_path: Path to complex Excel file with daily data
        """
        # Ensure daily file exists
        if not os.path.exists(daily_file_path):
            print(f"Error: Daily file not found at {daily_file_path}")
            return False

        try:
            # Extract player data from complex Excel format
            player_data = self._extract_player_data_from_complex_excel(daily_file_path)

            if player_data is None or len(player_data) == 0:
                print("Error: No valid player data could be extracted")
                return False

            # Read master file
            df_master = pd.read_excel(self.master_file_path)

            # Ensure master file has correct columns
            if "Username" not in df_master.columns:
                df_master["Username"] = []
            if "Streak" not in df_master.columns:
                df_master["Streak"] = []

            # Check for duplicate usernames in master file
            duplicate_users = df_master["Username"].duplicated().sum()
            if duplicate_users > 0:
                print(
                    f"WARNING: Found {duplicate_users} duplicate usernames in master file!"
                )
                print("Duplicated usernames:")
                for username in df_master[df_master["Username"].duplicated()][
                    "Username"
                ].unique():
                    print(f"  - {username}")

                # Remove duplicates, keeping the first occurrence
                df_master = df_master.drop_duplicates(subset=["Username"], keep="first")
                print(
                    "Duplicates have been removed from processing (first occurrence kept)"
                )

            # Create a set of all player usernames in today's data for quick lookup
            daily_usernames = {player["Username"] for player in player_data}

            # First pass to identify wheel winners
            wheel_winners = []

            # Process each player in daily data
            updates = []
            significant_lost_streaks = []  # Streaks of 4+ days
            minor_lost_streaks = []  # Streaks of 1-3 days
            new_players = []
            players_to_remove = []
            error_players = []  # Track players that cause errors
            history_updates = 0  # Counter for history updates

            # Process players in the daily data
            for player in player_data:
                try:
                    username = player["Username"]
                    hands = player["Hands"]

                    # Check if player exists in master file
                    player_rows = df_master[df_master["Username"] == username]
                    player_exists = len(player_rows) > 0

                    if player_exists:
                        # Get current streak
                        player_index = player_rows.index[0]
                        current_streak = df_master.at[player_index, "Streak"]

                        if hands >= self.hands_threshold:
                            # Increment streak
                            new_streak = current_streak + 1
                            df_master.at[player_index, "Streak"] = new_streak
                            updates.append(username)

                            # Check if player hit a wheel milestone
                            wheel_number = self._get_wheel_number(new_streak)
                            if wheel_number is not None:
                                # Player hit a new wheel milestone
                                wheel_winners.append(
                                    {
                                        "username": username,
                                        "streak": new_streak,
                                        "wheel_number": wheel_number,
                                    }
                                )
                                # Update player history with milestone message
                                status_message = f"Hit {new_streak} day milestone and earned Wheel {wheel_number} spin"
                                self._update_player_history(
                                    username, status_message, new_streak
                                )
                                history_updates += 1
                            else:
                                # Update player history with regular update
                                status_message = (
                                    f"Currently on a {new_streak} day streak"
                                )
                                self._update_player_history(
                                    username, status_message, new_streak
                                )
                                history_updates += 1
                        else:
                            # Record lost streak
                            status_message = f"Lost {current_streak} day streak (played {hands} hands)"
                            self._update_player_history(username, status_message, 0)
                            history_updates += 1

                            if current_streak >= 4:
                                significant_lost_streaks.append(
                                    f"{username} lost their {current_streak} day streak"
                                )
                            elif current_streak > 0:
                                minor_lost_streaks.append(username)

                            # Mark for removal from master sheet
                            players_to_remove.append(username)
                    else:
                        # New player, only add if they meet the threshold
                        if hands >= self.hands_threshold:
                            new_player = {
                                "Username": username,
                                "Streak": 1,
                            }
                            df_master = pd.concat(
                                [df_master, pd.DataFrame([new_player])],
                                ignore_index=True,
                            )
                            new_players.append(username)

                            # Update player history for new player
                            status_message = (
                                f"Joined streak system and started a 1 day streak"
                            )
                            self._update_player_history(username, status_message, 1)
                            history_updates += 1
                except Exception as e:
                    error_message = f"Error processing player '{username}': {str(e)}"
                    print(error_message)
                    error_players.append(error_message)
                    continue  # Skip to next player

            # Process players who are in the master sheet but NOT in today's data
            # These players automatically lose their streak
            for idx, row in df_master.iterrows():
                username = row["Username"]
                if username not in daily_usernames:
                    current_streak = row["Streak"]

                    # Update player history
                    status_message = (
                        f"Lost {current_streak} day streak due to inactivity"
                    )
                    self._update_player_history(username, status_message, 0)
                    history_updates += 1

                    if current_streak >= 4:
                        significant_lost_streaks.append(
                            f"{username} lost their {current_streak} day streak (not in today's data)"
                        )
                    elif current_streak > 0:
                        minor_lost_streaks.append(username)

                    # Mark for removal from master sheet
                    players_to_remove.append(username)

            # If we had any errors, show a summary
            if error_players:
                print(
                    f"\nWARNING: {len(error_players)} players caused errors and were skipped:"
                )
                for error in error_players[:5]:  # Show first 5 errors
                    print(f"  - {error}")
                if len(error_players) > 5:
                    print(f"  ... and {len(error_players) - 5} more")

            # Remove players with 0 streak
            if players_to_remove:
                df_master = df_master[~df_master["Username"].isin(players_to_remove)]

            # Generate unique wheel times for all winners
            if wheel_winners:
                try:
                    wheel_times = self._generate_unique_wheel_times(len(wheel_winners))

                    # Ensure we have enough times for all winners
                    while len(wheel_times) < len(wheel_winners):
                        wheel_times.append(f"TBD - Contact Admin")

                    # Assign times to winners
                    for i, winner in enumerate(wheel_winners):
                        winner["wheel_time"] = wheel_times[i]

                        # Update player history with wheel time
                        username = winner["username"]
                        streak = winner["streak"]
                        wheel_number = winner["wheel_number"]
                        wheel_time = winner["wheel_time"]

                        status_message = f"Hit {streak} day milestone and earned Wheel {wheel_number} spin scheduled for {wheel_time}"
                        self._update_player_history(username, status_message, streak)

                except Exception as e:
                    print(f"Error generating wheel times: {str(e)}")
                    # Fallback: assign "TBD" to all winners
                    for winner in wheel_winners:
                        winner["wheel_time"] = "TBD - Contact Admin"

                        # Update history with TBD time
                        username = winner["username"]
                        streak = winner["streak"]
                        wheel_number = winner["wheel_number"]

                        status_message = f"Hit {streak} day milestone and earned Wheel {wheel_number} spin (time TBD)"
                        self._update_player_history(username, status_message, streak)

            # Save updated master file
            df_master.to_excel(self.master_file_path, index=False)

            # Print wheel winners at the top
            if wheel_winners:
                print("\n🎡 WHEEL SPIN SCHEDULE TODAY 🎡")
                print("==============================")
                for winner in wheel_winners:
                    print(
                        f"{winner['username']} hit {winner['streak']} day streak, their Wheel {winner['wheel_number']} spin will be at {winner['wheel_time']} today"
                    )
                print("==============================\n")

            # Process referrals
            referral_bonuses = []
            try:
                # Create a dictionary of player -> hands for referral processing
                daily_players_hands = {
                    player["Username"]: player["Hands"] for player in player_data
                }

                # Initialize referral system
                referral_system = ReferralSystem(
                    "credentials.json", config.MASTER_SHEET_ID
                )

                # Update referral hands and check for milestones
                referral_bonuses = referral_system.update_hands_and_check_milestone(
                    daily_players_hands
                )

            except Exception as e:
                print(f"Warning: Error processing referrals: {e}")

            # Print referral bonuses if any
            if referral_bonuses:
                print("🎉 REFERRAL BONUSES TODAY 🎉")
                print("==============================")
                for bonus_message in referral_bonuses:
                    print(bonus_message)
                print("==============================\n")

            # Print summary
            print("--- PROCESSING COMPLETE ---")
            print(f"Total players processed: {len(player_data) - len(error_players)}")
            print(f"Updated Players: {len(updates)}")
            print(f"New Players: {len(new_players)}")
            print(
                f"Players that lost streaks: {len(significant_lost_streaks) + len(minor_lost_streaks)}"
            )
            print(f"Wheel winners: {len(wheel_winners)}")
            print(f"Referral bonuses: {len(referral_bonuses)}")
            print(f"History records updated: {history_updates}")

            # Only print notable lost streaks (4+ days)
            if significant_lost_streaks:
                print("\nSignificant Lost Streaks (4+ days):")
                for lost in significant_lost_streaks:
                    print(f"  - {lost}")

            return True

        except Exception as e:
            print(f"Error processing data: {str(e)}")
            import traceback

            traceback.print_exc()  # Print detailed error info
            return False


def lookup_player_command(tracker):
    """
    Command-line interface for player lookups
    """
    print("\n===== PLAYER LOOKUP =====")
    print("Enter player username to look up their streak status")
    username = input("Username: ")

    if not username.strip():
        print("No username entered. Exiting lookup.")
        return

    player_data = tracker.lookup_player(username.strip())

    if player_data is None:
        print(f"Player '{username}' not found in the history database.")
        return

    print("\n===== PLAYER STATUS =====")
    print(f"Username: {player_data['Username']}")
    print(f"Status: {player_data['LastUpdate']}")
    print(f"Last updated: {player_data['UpdateDate']}")
    print(f"Current streak: {player_data['CurrentStreak']}")
    print(f"Highest streak: {player_data['HighestStreak']}")
    print("========================")


def revive_streak_command(tracker):
    """
    Command-line interface for reviving a player's streak
    """
    print("\n===== REVIVE PLAYER STREAK (ADMIN FUNCTION) =====")
    username = input("Enter player username: ")

    if not username.strip():
        print("No username entered. Exiting.")
        return

    try:
        new_streak = int(input("Enter streak value to set: "))
        if new_streak < 1:
            print("Error: Streak value must be positive")
            return
    except ValueError:
        print("Error: Please enter a valid number")
        return

    # Try to revive streak
    result, message = tracker.revive_player_streak(username.strip(), new_streak)

    # Handle confirmation case
    if result == "confirm":
        print(f"\nWARNING: {message}")
        confirm = input("Do you want to proceed? (y/n): ").lower()

        if confirm == "y" or confirm == "yes":
            result, message = tracker.confirm_revive_streak(
                username.strip(), new_streak
            )
            print(f"\n{message}")
        else:
            print("\nOperation cancelled.")
    else:
        print(f"\n{message}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Process poker hand data and track streaks."
    )
    parser.add_argument("--daily", help="Path to daily Excel file with player data")
    parser.add_argument(
        "--master",
        default="data/master_streak.xlsx",
        help="Path to master Excel file (default: data/master_streak.xlsx)",
    )
    parser.add_argument(
        "--history",
        help="Path to history Excel file (default: derived from master path)",
    )
    parser.add_argument(
        "--threshold",
        type=int,
        default=100,
        help="Minimum hands to count as streak day (default: 100)",
    )
    parser.add_argument(
        "--lookup", action="store_true", help="Look up a player's streak status"
    )
    parser.add_argument(
        "--revive",
        action="store_true",
        help="Revive or set a player's streak (admin function)",
    )
    parser.add_argument("--username", help="Username for revive function")
    parser.add_argument("--streak", type=int, help="Streak value for revive function")

    args = parser.parse_args()

    # Ensure master directory exists
    master_dir = os.path.dirname(args.master)
    if master_dir and not os.path.exists(master_dir):
        os.makedirs(master_dir)

    # If history path specified, ensure its directory exists
    if args.history:
        history_dir = os.path.dirname(args.history)
        if history_dir and not os.path.exists(history_dir):
            os.makedirs(history_dir)

    # Initialize tracker
    tracker = PokerStreakTracker(args.master, args.threshold, args.history)

    # Handle revive request
    if args.revive:
        if args.username and args.streak is not None:
            result, message = tracker.revive_player_streak(args.username, args.streak)

            if result == "confirm":
                print(f"\nWARNING: {message}")
                confirm = input("Do you want to proceed? (y/n): ").lower()

                if confirm == "y" or confirm == "yes":
                    result, message = tracker.confirm_revive_streak(
                        args.username, args.streak
                    )
                    print(f"\n{message}")
                else:
                    print("\nOperation cancelled.")
            else:
                print(f"\n{message}")
        else:
            revive_streak_command(tracker)
    # Handle lookup request
    elif args.lookup:
        lookup_player_command(tracker)
    # Process daily data if provided
    elif args.daily:
        if not os.path.exists(args.daily):
            print(f"Error: Daily file not found at {args.daily}")
        else:
            tracker.process_daily_data(args.daily)
    # If neither, show options
    else:
        print("\nPoker Streak Tracker Options:")
        print("1) Process daily data")
        print("2) Look up player status")
        print("3) Revive player streak (Admin)")
        print("4) Exit")

        choice = input("\nSelect an option (1-4): ")

        if choice == "1":
            file_path = input("Enter path to daily Excel file: ")
            if os.path.exists(file_path):
                tracker.process_daily_data(file_path)
            else:
                print(f"Error: File not found at {file_path}")
        elif choice == "2":
            lookup_player_command(tracker)
        elif choice == "3":
            revive_streak_command(tracker)
        else:
            print("Exiting...")
